import pandas as pd
from openpyxl import load_workbook

def analyze_customer_data(file_path: str) -> None:
    df = pd.read_excel(file_path, sheet_name="data")
    total_customers = len(df)  
    average_age = df["Age"].mean()  
    total_purchase = df["PurchaseAmount"].sum()  
    top_spender = df.loc[df["PurchaseAmount"].idxmax(), "Name"]  
    results = [
        ["Metric", "Value"],
        ["Total Customers", total_customers],
        ["Average Age", average_age],
        ["Total Purchase Amount", total_purchase],
        ["Top Spender", top_spender]
    ]
    book = load_workbook(file_path)
    if "result" in book.sheetnames:
        book.remove(book["result"]) 
    result_ws = book.create_sheet("result") 
    for row in results:
        result_ws.append(row)
    book.save(file_path)
    return None



if __name__ == "__main__":
    analyze_customer_data(r'C:\AI\1. Data Exploration and Preprocessing\1. Understanding dataset structure, sample records\sampledata.xlsx')